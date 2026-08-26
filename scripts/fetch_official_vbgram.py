#!/usr/bin/env python3
"""Fetch VB-G RAM G official MIS data using a real browser.

Goal: obtain the same Daily Report workbook (RepDay + Sheet1 + VBG) without
manual upload. The script is defensive: it never publishes partial/bad data.

Optional GitHub secrets/env:
  VBGRAM_DAILY_REPORT_URL : direct report page URL, if known
  VBGRAM_USERNAME / VBGRAM_PASSWORD : only if official site prompts login
  VBGRAM_COOKIE : raw Cookie header if a session cookie is required

The committed defaults are the official links already shown in the portal.
"""
import os, re, json, shutil, sys, time
from pathlib import Path
from datetime import datetime, timezone

ROOT=Path(__file__).resolve().parents[1]
INCOMING=ROOT/'incoming'; RAW=ROOT/'raw'; DATA=ROOT/'data'
INCOMING.mkdir(exist_ok=True); RAW.mkdir(exist_ok=True); DATA.mkdir(exist_ok=True)

HOME='https://vbgramg.dord.gov.in/vbgramg/home.aspx'
MIS='https://vbgramgrep.dord.gov.in/VBGRAMG/MISreport.aspx'
ONGOING='https://vbgramgrep.dord.gov.in/VBGRAMG/dynamic_work_details.aspx?payload=4PmH2eRA9khYNUNqz1h5yt9D8POKLA7Afp0nercX3xt22K65u-hNco55SZiMHr78IufQr-Pyxw1-2tJEz-65UMtG5kOTBzCEHurJmRrAtoAIfVSTK-qhJdX02vLZMWrVbwM-oS9xX58g6SiO5ODhhFid9RqKvnwTnS-hLkXfa1-25phIp66JlphIcilUU7cK'
REPORT_URL=os.environ.get('VBGRAM_DAILY_REPORT_URL','').strip() or MIS
USERNAME=os.environ.get('VBGRAM_USERNAME','').strip()
PASSWORD=os.environ.get('VBGRAM_PASSWORD','').strip()
COOKIE=os.environ.get('VBGRAM_COOKIE','').strip()

status={'startedAt':datetime.now(timezone.utc).isoformat(),'ok':False,'steps':[],'source':'Official VB-G RAM G'}
def note(step,ok=True,detail=''):
    status['steps'].append({'step':step,'ok':bool(ok),'detail':str(detail)[:1000]})
    print(('OK ' if ok else 'WARN ')+step+(' :: '+str(detail) if detail else ''))

def save_status():
    status['finishedAt']=datetime.now(timezone.utc).isoformat()
    (DATA/'fetch-status.json').write_text(json.dumps(status,ensure_ascii=False,indent=2),encoding='utf-8')
    (ROOT/'auto-status.js').write_text('window.AUTO_FETCH_STATUS='+json.dumps(status,ensure_ascii=False,separators=(',',':'))+';\n',encoding='utf-8')

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
except Exception as e:
    note('playwright import',False,e); save_status(); raise

def maybe_login(page):
    # Generic login support only when credentials are supplied.
    if not (USERNAME and PASSWORD): return False
    pw=page.locator('input[type="password"]')
    if pw.count()==0: return False
    user=page.locator('input[type="text"],input[type="email"]')
    if user.count(): user.first.fill(USERNAME)
    pw.first.fill(PASSWORD)
    btn=page.get_by_role('button',name=re.compile('login|sign in|submit|प्रवेश',re.I))
    if btn.count(): btn.first.click()
    else: pw.first.press('Enter')
    page.wait_for_timeout(2000)
    note('login submitted',True)
    return True

def table_dump(page,label):
    try:
        tables=page.locator('table')
        out=[]
        for i in range(min(tables.count(),60)):
            t=tables.nth(i)
            rows=[]
            trs=t.locator('tr')
            for r in range(min(trs.count(),3000)):
                cells=trs.nth(r).locator('th,td')
                rows.append([cells.nth(c).inner_text().strip() for c in range(cells.count())])
            if rows: out.append(rows)
        (RAW/f'{label}_tables.json').write_text(json.dumps(out,ensure_ascii=False),encoding='utf-8')
        note(f'{label} tables captured',True,len(out))
    except Exception as e: note(f'{label} tables captured',False,e)

def try_download(page):
    """Try common Excel/export controls; return downloaded path or None."""
    # buttons/links with likely export text
    pats=re.compile(r'excel|xlsx|download|export|डाउनलोड|एक्सेल',re.I)
    candidates=page.get_by_role('link',name=pats)
    if candidates.count()==0: candidates=page.get_by_role('button',name=pats)
    for i in range(min(candidates.count(),25)):
        el=candidates.nth(i)
        try:
            if not el.is_visible(): continue
            with page.expect_download(timeout=12000) as di:
                el.click()
            d=di.value
            suggested=d.suggested_filename or 'Daily Report.xlsx'
            if not suggested.lower().endswith(('.xlsx','.xls','.csv')): suggested+=' .xlsx'
            target=RAW/suggested.replace('/','_')
            d.save_as(str(target))
            note('official export downloaded',True,target.name)
            return target
        except Exception:
            continue
    return None

def validate_workbook(path):
    if not path or path.suffix.lower() not in ('.xlsx','.xlsm','.xls'): return False
    try:
        from openpyxl import load_workbook
        wb=load_workbook(path,read_only=True,data_only=True)
        need={'RepDay','Sheet1','VBG'}
        ok=need.issubset(set(wb.sheetnames))
        note('workbook sheet validation',ok,','.join(wb.sheetnames[:20]))
        wb.close(); return ok
    except Exception as e:
        note('workbook sheet validation',False,e); return False



def parse_official_summary_from_tables():
    """Parse the 8-Janpad official summary directly from captured MIS HTML tables.
    This is the fallback when the official portal does not expose a downloadable workbook.
    """
    source=RAW/'mis_tables.json'
    if not source.exists(): return None
    try: tables=json.loads(source.read_text(encoding='utf-8'))
    except Exception as e:
        note('official summary table parse',False,e); return None
    valid={'AMARPATAN','MAIHAR','RAMNAGAR','MAJHGAWAN','NAGOD','RAMPUR BAGHELAN','SATNA','UNCHAHARA'}
    fields=['totalGP','musterGP','dysfunctionalGP','labourAll','mrAll','ongoingAll','labourIndividual','mrIndividual','labourCommunity','mrCommunity','pmayOngoing','pmayMR','ekLabour','ekOngoing','ekMR']
    idx=[2,3,4,6,7,8,10,11,12,13,15,16,18,19,20]
    def n(v):
        try:return float(str(v).replace(',','').strip() or 0)
        except:return 0.0
    best={}
    for table in tables:
        found={}
        for r in table:
            if len(r)<21: continue
            jan=str(r[1]).strip().upper()
            if jan not in valid: continue
            z={'janpad':jan}
            for k,i in zip(fields,idx): z[k]=n(r[i])
            if z['totalGP']>0 and z['ongoingAll']>0: found[jan]=z
        if len(found)>len(best): best=found
    if set(best)!=valid:
        note('official summary table parse',False,f'Expected 8 Janpads, got {len(best)}'); return None
    import csv
    out=DATA/'official-summary.csv'
    with out.open('w',encoding='utf-8-sig',newline='') as f:
        w=csv.DictWriter(f,fieldnames=['janpad']+fields); w.writeheader()
        for j in ['AMARPATAN','MAIHAR','RAMNAGAR','MAJHGAWAN','NAGOD','RAMPUR BAGHELAN','SATNA','UNCHAHARA']:w.writerow(best[j])
    note('official summary table parse',True,'8 Janpads -> data/official-summary.csv')
    return out

with sync_playwright() as p:
    browser=p.chromium.launch(headless=True,args=['--no-sandbox'])
    context=browser.new_context(accept_downloads=True,viewport={'width':1600,'height':1000})
    if COOKIE:
        # Parse simple Cookie header into cookies for both official hosts.
        cookies=[]
        for pair in COOKIE.split(';'):
            if '=' not in pair: continue
            k,v=pair.strip().split('=',1)
            for domain in ['vbgramgrep.dord.gov.in','vbgramg.dord.gov.in']:
                cookies.append({'name':k,'value':v,'domain':domain,'path':'/'})
        try: context.add_cookies(cookies); note('session cookie loaded',True,len(cookies))
        except Exception as e: note('session cookie loaded',False,e)
    page=context.new_page()

    # Home warms cookies/session.
    try:
        r=page.goto(HOME,wait_until='domcontentloaded',timeout=60000)
        note('Home opened',bool(r and r.ok),getattr(r,'status',None))
    except Exception as e: note('Home opened',False,e)

    # MIS page: capture tables and try official Excel export.
    downloaded=None
    for url,label in [(REPORT_URL,'mis'),(ONGOING,'ongoing')]:
        try:
            r=page.goto(url,wait_until='domcontentloaded',timeout=90000)
            note(f'{label} opened',bool(r and r.ok),getattr(r,'status',None))
            maybe_login(page)
            page.wait_for_timeout(2500)
            table_dump(page,label)
            html=page.content(); (RAW/f'{label}.html').write_text(html,encoding='utf-8')
            if not downloaded: downloaded=try_download(page)
        except Exception as e:
            note(f'{label} opened',False,e)

    # Search any captured browser responses/downloads is not needed when export worked.
    browser.close()

# Prefer a valid official workbook. If unavailable, fall back to the live official 8-Janpad HTML table.
if downloaded and validate_workbook(downloaded):
    dest=INCOMING/'Daily Report.xlsx'
    shutil.copy2(downloaded,dest)
    status['ok']=True; status['updateMode']='workbook'; status['workbook']=str(dest.relative_to(ROOT))
    note('fresh Daily Report installed',True,dest)
else:
    summary=parse_official_summary_from_tables()
    if summary:
        status['ok']=True; status['updateMode']='summary'; status['summary']=str(summary.relative_to(ROOT))
        note('summary fallback enabled',True,'Official cards/table can refresh without full Excel export')
    else:
        status['ok']=False
        note('fresh official data installed',False,'Neither a valid workbook nor the 8-Janpad official summary table was detected. Previous valid report remains live.')

save_status()
if not status['ok']:
    sys.exit(2)
