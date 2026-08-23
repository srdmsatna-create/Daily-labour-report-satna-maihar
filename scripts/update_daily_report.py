#!/usr/bin/env python3
import os, json, re, urllib.request, tempfile
from pathlib import Path
from datetime import datetime, timezone
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
OUT=ROOT/'auto-data.js'
INCOMING=ROOT/'incoming'/'Daily Report.xlsx'
ROOT_XLSX=ROOT/'Daily Report.xlsx'
URL=os.environ.get('DAILY_REPORT_XLSX_URL','').strip()
VALID_JANPADS={'AMARPATAN','MAIHAR','RAMNAGAR','MAJHGAWAN','NAGOD','RAMPUR BAGHELAN','SATNA','UNCHAHARA'}

def n(v):
    try:return float(v or 0)
    except:return 0.0

def c(v): return '' if v is None else str(v).strip()
def up(v): return c(v).upper()
def janpad(v):
    j=up(v)
    return 'SATNA' if j=='SOHAWAL' else j

def obtain():
    if URL:
        fd,name=tempfile.mkstemp(suffix='.xlsx'); os.close(fd)
        req=urllib.request.Request(URL,headers={'User-Agent':'Mozilla/5.0 DailyReportBot/1.1'})
        with urllib.request.urlopen(req,timeout=90) as r, open(name,'wb') as f:f.write(r.read())
        return Path(name),URL
    if INCOMING.exists(): return INCOMING,'repo:incoming/Daily Report.xlsx'
    if ROOT_XLSX.exists(): return ROOT_XLSX,'repo:Daily Report.xlsx'
    raise SystemExit('No source workbook. Set DAILY_REPORT_XLSX_URL secret or add incoming/Daily Report.xlsx')

def extract_date(text):
    m=re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})',c(text))
    return f'{int(m.group(1)):02d}-{int(m.group(2)):02d}-{m.group(3)}' if m else None

def parse(path):
    wb=load_workbook(path,data_only=True,read_only=True)
    for sh in ('RepDay','Sheet1','VBG'):
        if sh not in wb.sheetnames: raise SystemExit(f'Missing sheet: {sh}')

    # RepDay = Engineer / Cluster / GP drill-down only.
    rep=wb['RepDay']; repvals=list(rep.iter_rows(values_only=True))
    title=c(repvals[0][0]) if repvals else ''
    rep_date=extract_date(title); rows=[]; gpmap={}
    for r in repvals[3:]:
        if len(r)<12 or not c(r[0]) or not c(r[1]) or not c(r[5]): continue
        z={'janpad':janpad(r[0]),'engineer':c(r[1]),'cluster':c(r[2]),'ongoing':n(r[3]),'panchayat':up(r[5]),'gps':n(r[6]),'gpsProgress':n(r[7]),'labour':n(r[8]),'worksMR':n(r[9]),'noEkyc':n(r[10]),'mrs':n(r[11])}
        rows.append(z); gpmap[(z['janpad'],z['panchayat'])]=(z['engineer'],z['cluster'])
    if not rows: raise SystemExit('RepDay has no usable GP rows')

    # Sheet1 upper table = official work-load/category metrics (including total ongoing works).
    s1=wb['Sheet1']; svals=list(s1.iter_rows(values_only=True))
    sheet1_title=c(svals[0][0]) if svals else ''
    sheet1_date=extract_date(sheet1_title); official=[]
    for idx in (3,4,5,7,8,9,10,11):  # Excel rows 4,5,6,8..12
        if idx>=len(svals): continue
        r=svals[idx]
        if len(r)<20 or not c(r[1]): continue
        official.append({'janpad':janpad(r[1]),'totalGP':n(r[2]),'musterGP':n(r[3]),'dysfunctionalGP':n(r[4]),'labourAll':n(r[5]),'mrAll':n(r[6]),'ongoingAll':n(r[7]),'labourIndividual':n(r[9]),'mrIndividual':n(r[10]),'labourCommunity':n(r[11]),'mrCommunity':n(r[12]),'pmayOngoing':n(r[14]),'pmayMR':n(r[15]),'ekLabour':n(r[17]),'ekOngoing':n(r[18]),'ekMR':n(r[19])})

    # Sheet1 lower embedded table = Screen 2 authority.
    # Columns: GP, GP with progress, labour, works with MR, no-eKYC, Muster Rolls.
    header_idx=None
    for i,r in enumerate(svals):
        if len(r)>2 and 'total no. of gram panchayats' in c(r[2]).lower():
            header_idx=i; break
    daily=[]
    if header_idx is not None:
        for r in svals[header_idx+2:]:
            if len(r)<8: continue
            jan=janpad(r[1])
            if jan not in VALID_JANPADS: continue
            daily.append({'janpad':jan,'totalGP':n(r[2]),'gpsProgress':n(r[3]),'labour':n(r[4]),'worksMR':n(r[5]),'noEkyc':n(r[6]),'mrs':n(r[7])})
    if len(daily)!=8:
        raise SystemExit(f'Screen 2 table parse failed: expected 8 Janpads, got {len(daily)}')

    # VBG = work-level add-ons for engineer cards, category-wise work and expenditure buckets.
    v=wb['VBG']; wm={}; cm={}
    inst=re.compile(r'(school|prathmik|madhyamik|shala|vidyalaya|प्राथमिक|माध्यमिक|शाला|विद्यालय)',re.I)
    def exp_bucket(p):
        if p<=0: return 'b0'
        if p<=25: return 'b25'
        if p<=60: return 'b60'
        if p<=75: return 'b75'
        if p<=90: return 'b90'
        return 'b90p'
    for i,r in enumerate(v.iter_rows(values_only=True),start=1):
        if i<5 or len(r)<22: continue
        jan,gp,fy,status,code,name,wt=janpad(r[2]),up(r[3]),c(r[4]),c(r[5]).lower(),c(r[6]),c(r[7]),c(r[8])
        if not jan or not gp or not code or 'ongoing' not in status: continue
        eng,cl=gpmap.get((jan,gp),('Unmapped','Unmapped')); key=(jan,eng,cl,gp)
        z=wm.setdefault(key,{'janpad':jan,'engineer':eng,'cluster':cl,'panchayat':gp,'workTotal':0,'pmayOngoing':0,'ekOngoing':0,'currentFYActive':0})
        z['workTotal']+=1
        if 'pmay' in wt.lower(): z['pmayOngoing']+=1
        if wt.lower()=='ek bagiya' and fy in ('2025-2026','2026-2027') and not inst.search(name): z['ekOngoing']+=1
        if n(r[21])>0: z['currentFYActive']+=1

        category=wt or 'Unspecified'
        sanctioned=n(r[11]) if len(r)>11 else (n(r[9])+n(r[10]))
        booked=(n(r[16])+n(r[17])) if len(r)>17 else 0
        ep=(booked*100/sanctioned) if sanctioned>0 else 0
        ck=(jan,eng,cl,gp,category)
        q=cm.setdefault(ck,{'janpad':jan,'engineer':eng,'cluster':cl,'panchayat':gp,'category':category,'workCount':0,'totalSanction':0,'totalBooked':0,'b0':0,'b25':0,'b60':0,'b75':0,'b90':0,'b90p':0})
        q['workCount']+=1; q['totalSanction']+=sanctioned; q['totalBooked']+=booked; q[exp_bucket(ep)]+=1

    return {'title':title,'rows':rows,'official':official,'daily':daily,'workmix':list(wm.values()),'categorymix':list(cm.values()),'_sourceDates':{'RepDay':rep_date,'Sheet1':sheet1_date}}

path,source=obtain(); data=parse(path)
dates=data.pop('_sourceDates',{})
data['meta']={'mode':'auto','status':'ok','updatedAt':datetime.now(timezone.utc).isoformat(),'source':source,'rowCount':len(data['rows']),'sourceDates':dates}
OUT.write_text('window.AUTO_REPORT='+json.dumps(data,ensure_ascii=False,separators=(',',':'))+';\n',encoding='utf-8')
print(f'Wrote {OUT} with {len(data["rows"])} GP rows and {len(data["daily"])} Screen-2 Janpad rows')
